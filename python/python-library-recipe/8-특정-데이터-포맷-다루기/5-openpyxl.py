import openpyxl  # pip install openpyxl  


"""파이썬에서 Microsoft Excel 파일의 읽기나 쓰기 등을 실행하는 기능을 제공하는 openpyxl 패키지
opnepyxl은 Microsoft Office 2007 이후의 xlsx/xlsm/xltx/xltm 포맷 지원
"""
def main():
    """엑셀 파일 읽어오기 """
    path = './python/python-library-recipe/8-특정-데이터-포맷-다루기/sample.xlsx'
    # 엑셀 파일을 읽기
    wb = openpyxl.load_workbook(path)  # Workbook 객체 반환
    # 시트 이름 리스트 얻기
    print(wb.sheetnames)
    # 시트를 이름으로 지정하여 얻기
    ws = wb['Sheet 1']  # Worksheet객체
    print(ws.max_column)
    print(ws.max_row)
    print(ws['A4'].value)
    # 셀 값을 얻기
    a2 = ws.cell(row=2, column=1)  # A2를 지정하려면 row=2, column=1 (row, column을 사용할 때 맨 처음이 0이 아닌 1임)
    print(a2.value)
    b5 = ws.cell(row=5, column=2)
    print(b5.value)  # load_workbook() 함수의 data_only 기본값 False이므로, 수식 출력

    # 셀 값을 순서대로 얻기
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet 1']

    print('A1 -> A2 -> ... -> B1 -> B2의 순서대로 값을 얻습니다.')
    for row in ws.rows:
        for cell in row:
            print(cell.value)

    print('A1 -> B1 -> A2 -> ...의 순서대로 값을 얻습니다.')
    for column in ws.columns:
        for cell in column:
            print(cell.value)
    
    """엑셀 파일 쓰기"""
    wb = openpyxl.Workbook()
    # 엑셀 파일 시트를 삽입
    # 인수: title, index(삽입할 위치, 0이면 가장 왼쪽에 삽입)
    # 반환: Worksheet 객체
    ws = wb.create_sheet(title='New Sheet', index=0)
    ws['A1'] = 100
    # 엑셀 파일을 저장
    wb.save(filename='new_book.xlsx')


if __name__ == "__main__":
    main()
